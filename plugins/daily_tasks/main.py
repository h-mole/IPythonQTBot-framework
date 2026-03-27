"""
每日任务提醒器插件 - 提供任务管理和提醒功能
"""

import os
import sys
import json
from pathlib import Path
from datetime import datetime, time as dt_time
from PySide6.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QSplitter,
    QPushButton,
    QLabel,
    QFrame,
    QGroupBox,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QComboBox,
    QFileDialog,
    QMessageBox,
    QDialog,
    QLineEdit,
    QTextEdit,
    QDateTimeEdit,
    QFormLayout,
    QDialogButtonBox,
    QCheckBox,
    QTimeEdit,
    QListWidget,
    QListWidgetItem,
    QAbstractItemView,
    QMenu,
)
from PySide6.QtCore import Qt, QTimer, QTime, Signal
from PySide6.QtGui import QFont, QAction, QColor
from PySide6.QtCore import QDateTime as QtCore_QDateTime
from app_qt.plugin_manager import PluginManager, exec_main_thread_callback
import openpyxl
from openpyxl.utils import get_column_letter
import plyer
from app_qt.configs import PLUGIN_DATA_DIR
from app_qt.widgets.custom_checkbox import CustomCheckBox
from app_qt.plugin_i18n import PluginI18n
from .colors import TaskColorManager, DateUrgency

# Initialize plugin i18n
import os as _os
_plugin_dir = _os.path.dirname(_os.path.abspath(__file__))
_plugin_i18n = PluginI18n("daily_tasks", Path(_plugin_dir))
_ = _plugin_i18n.gettext
# 默认分类（使用英文存储，显示时翻译）
DEFAULT_CATEGORIES = ["Paper", "Project"]
DEFAULT_SUBCATEGORIES = ["Admin", "Project", "Meeting", "Study"]
DEFAULT_STATUSES = ["Completed", "In Progress", "Cancelled", "Not Started"]
REMINDER_TYPES = ["Daily", "On Due Date", "Once"]

# 显示翻译映射
def tr_category(cat):
    """翻译分类显示"""
    mapping = {
        "Paper": _("Paper"),
        "Project": _("Project"),
        "Admin": _("Admin"),
        "Meeting": _("Meeting"),
        "Study": _("Study"),
    }
    return mapping.get(cat, cat)

def tr_status(status):
    """翻译状态显示"""
    mapping = {
        "Completed": _("Completed"),
        "In Progress": _("In Progress"),
        "Cancelled": _("Cancelled"),
        "Not Started": _("Not Started"),
    }
    return mapping.get(status, status)

def tr_reminder_type(rt):
    """翻译提醒方式显示"""
    mapping = {
        "Daily": _("Daily"),
        "On Due Date": _("On Due Date"),
        "Once": _("Once"),
    }
    return mapping.get(rt, rt)

# 插件数据文件路径
DAILY_TASKS_DATA_DIR = os.path.join(PLUGIN_DATA_DIR, "daily_tasks")

TASKS_FILE = os.path.join(DAILY_TASKS_DATA_DIR, "daily_tasks.xlsx")
CATEGORIES_FILE = os.path.join(DAILY_TASKS_DATA_DIR, "task_categories.json")

# 特殊日期映射

SPECIAL_DATES = [
    ("1999-09-10", _("Long Term"), "long_term"),
    ("1999-09-09", _("No Deadline"), "no_deadline"),
]

class MultiSelectFilter(QWidget):
    """多选筛选器组件"""

    filterChanged = Signal()

    def __init__(self, label_text, items=None, parent=None, item_display_mapper=None):
        super().__init__(parent)
        self.items = items or []
        self.all_option = _("All")
        self.item_display_mapper = item_display_mapper  # 可选的显示文本映射函数

        # 创建布局
        layout = QHBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(5)
        self.setLayout(layout)

        # 标签
        layout.addWidget(QLabel(label_text))

        # 筛选按钮
        self.filter_btn = QPushButton(_("Please select..."))
        self.filter_btn.setMinimumWidth(120)
        self.filter_btn.clicked.connect(self.show_menu)
        layout.addWidget(self.filter_btn)

        # 选中状态记录
        self.selected_items = set()
        self.select_all = True  # 默认全选

        # 更新按钮文本
        self.update_button_text()

    def show_menu(self):
        """显示选择菜单"""
        menu = QMenu(self)

        # "全部"选项
        all_action = menu.addAction(self.all_option)
        all_action.setCheckable(True)
        all_action.setChecked(self.select_all)
        all_action.triggered.connect(lambda: self.toggle_all())
        menu.addSeparator()

        # 其他选项
        self.item_actions = {}
        for item in self.items:
            # 使用映射函数翻译显示文本（如果有）
            display_text = self.item_display_mapper(item) if self.item_display_mapper else item
            action = menu.addAction(display_text)
            action.setCheckable(True)
            # 如果是全选模式，所有项都勾选
            is_selected = self.select_all or item in self.selected_items
            action.setChecked(is_selected)
            self.item_actions[item] = action
            action.triggered.connect(lambda checked, name=item: self.toggle_item(name))

        # 显示菜单
        menu.exec_(self.filter_btn.mapToGlobal(self.filter_btn.rect().bottomLeft()))

    def toggle_all(self):
        """切换全选状态"""
        self.select_all = not self.select_all
        if self.select_all:
            self.selected_items.clear()
        self.update_button_text()
        self.filterChanged.emit()

    def toggle_item(self, item_name):
        """切换单项选择状态"""
        if self.select_all:
            # 如果当前是全选，取消某个项
            self.select_all = False
            # 初始化选中集合为除了当前点击项之外的所有项
            self.selected_items = set(self.items) - {item_name}
        else:
            # 如果当前不是全选，切换该项
            if item_name in self.selected_items:
                self.selected_items.remove(item_name)
            else:
                self.selected_items.add(item_name)

            # 检查是否变成了全选
            if len(self.selected_items) == len(self.items):
                self.select_all = True
                self.selected_items.clear()

        self.update_button_text()
        self.filterChanged.emit()

    def update_button_text(self):
        """更新按钮显示文本"""
        if self.select_all:
            self.filter_btn.setText(_("All ({})").format(len(self.items)))
        elif len(self.selected_items) == 0:
            self.filter_btn.setText(_("No items selected"))
        elif len(self.selected_items) <= 3:
            # 使用映射函数翻译显示文本（如果有）
            if self.item_display_mapper:
                display_items = [self.item_display_mapper(item) for item in sorted(self.selected_items)]
            else:
                display_items = sorted(self.selected_items)
            self.filter_btn.setText(", ".join(display_items))
        else:
            self.filter_btn.setText(_("Selected {} items").format(len(self.selected_items)))

    def get_selected(self):
        """获取选中的项目"""
        if self.select_all:
            return set(self.items)
        return self.selected_items.copy()

    def is_select_all(self):
        """是否全选"""
        return self.select_all

    def set_items(self, items):
        """设置筛选项"""
        self.items = items
        if self.select_all:
            self.selected_items.clear()
        self.update_button_text()

    def set_selected_items(self, items_to_select):
        """设置选中的项目"""
        self.selected_items = set(items_to_select)
        self.select_all = len(self.selected_items) == len(self.items)
        self.update_button_text()
        self.filterChanged.emit()


class TaskDialog(QDialog):
    """任务编辑对话框"""

    def __init__(self, parent=None, task_data=None):
        super().__init__(parent)
        self.task_data = task_data

        # 加载用户自定义分类
        self.load_user_categories()

        self.init_ui()

    def load_user_categories(self):
        """加载用户自定义的分类"""
        try:
            if os.path.exists(CATEGORIES_FILE):
                with open(CATEGORIES_FILE, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    self.user_categories = data.get("categories", DEFAULT_CATEGORIES)
                    self.user_subcategories = data.get(
                        "subcategories", DEFAULT_SUBCATEGORIES
                    )
            else:
                self.user_categories = DEFAULT_CATEGORIES
                self.user_subcategories = DEFAULT_SUBCATEGORIES
        except Exception as e:
            print(f"Error loading categories: {e}")
            self.user_categories = DEFAULT_CATEGORIES
            self.user_subcategories = DEFAULT_SUBCATEGORIES

    def init_ui(self):
        """初始化界面"""
        self.setWindowTitle(_("Add/Edit Task"))
        self.setMinimumWidth(500)

        layout = QFormLayout()
        self.setLayout(layout)

        # 任务大类
        self.category_combo = QComboBox()
        self.category_combo.setEditable(True)
        self.category_combo.addItems(self.user_categories)
        if self.task_data and self.task_data.get("category"):
            self.category_combo.setCurrentText(self.task_data["category"])
        layout.addRow(_("Category:"), self.category_combo)

        # 任务小类
        self.subcategory_combo = QComboBox()
        self.subcategory_combo.setEditable(True)
        self.subcategory_combo.addItems(self.user_subcategories)
        if self.task_data and self.task_data.get("subcategory"):
            self.subcategory_combo.setCurrentText(self.task_data["subcategory"])
        layout.addRow(_("Subcategory:"), self.subcategory_combo)

        # 任务说明
        self.description_edit = QTextEdit()
        self.description_edit.setMaximumHeight(100)
        if self.task_data and self.task_data.get("description"):
            self.description_edit.setPlainText(self.task_data["description"])
        layout.addRow(_("Description:"), self.description_edit)

        # 完成日期
        self.due_date_edit = QDateTimeEdit()
        self.due_date_edit.setCalendarPopup(True)
        self.due_date_edit.setDisplayFormat("yyyy-MM-dd")
        if self.task_data and self.task_data.get("due_date"):
            try:
                due_date = datetime.strptime(self.task_data["due_date"], "%Y-%m-%d")
                qt_datetime = QtCore_QDateTime(
                    due_date.year, due_date.month, due_date.day, 0, 0, 0  # 时间设置为 0
                )
                self.due_date_edit.setDateTime(qt_datetime)
            except:
                self.due_date_edit.setDateTime(QtCore_QDateTime.currentDateTime())
        else:
            self.due_date_edit.setDateTime(QtCore_QDateTime.currentDateTime())
        layout.addRow(_("Due Date:"), self.due_date_edit)

        # 提醒方式
        self.reminder_type_combo = QComboBox()
        self.reminder_type_combo.addItems(REMINDER_TYPES)
        if self.task_data and self.task_data.get("reminder_type"):
            self.reminder_type_combo.setCurrentText(self.task_data["reminder_type"])
        layout.addRow(_("Reminder Type:"), self.reminder_type_combo)

        # 提醒时间（只设置时刻，不设置日期）
        self.reminder_time_edit = QTimeEdit()
        self.reminder_time_edit.setDisplayFormat("HH:mm")

        # 特殊时间下拉菜单
        self.special_time_combo = QComboBox()
        self.special_time_combo.addItem(_("Normal Date"), "normal")
        for date_str, text, optionname in SPECIAL_DATES:
            self.special_time_combo.addItem(_(text), optionname)
        self.special_time_combo.currentTextChanged.connect(self.on_special_time_changed)

        # 创建水平布局放置时间选择器和下拉菜单
        time_layout = QHBoxLayout()
        time_layout.setSpacing(5)
        time_layout.addWidget(self.reminder_time_edit)
        time_layout.addWidget(self.special_time_combo)

        if self.task_data and self.task_data.get("reminder_time"):
            try:
                # 尝试从旧格式解析（包含日期）
                reminder_time = datetime.strptime(
                    self.task_data["reminder_time"], "%Y-%m-%d %H:%M"
                )
                time_obj = QTime(reminder_time.hour, reminder_time.minute)
                self.reminder_time_edit.setTime(time_obj)
                # 检查是否是特殊日期
                for date_str, text, optionname in SPECIAL_DATES:
                    if date_str in self.task_data.get("due_date", ""):
                        self.special_time_combo.setCurrentText(text)
                        self.reminder_time_edit.setEnabled(False)
                        break
            except:
                try:
                    # 尝试从新格式解析（只有时间）
                    reminder_time = datetime.strptime(
                        self.task_data["reminder_time"], "%H:%M"
                    )
                    time_obj = QTime(reminder_time.hour, reminder_time.minute)
                    self.reminder_time_edit.setTime(time_obj)
                except:
                    self.reminder_time_edit.setTime(QTime.fromString("10:30", "HH:mm"))
        else:
            self.reminder_time_edit.setTime(QTime.currentTime())

        layout.addRow(_("Reminder Time:"), time_layout)

        # 完成状态
        self.status_combo = QComboBox()
        self.status_combo.addItems(DEFAULT_STATUSES)
        if self.task_data and self.task_data.get("status"):
            self.status_combo.setCurrentText(self.task_data["status"])
        layout.addRow(_("Status:"), self.status_combo)

        # 备注
        self.notes_edit = QTextEdit()
        self.notes_edit.setMaximumHeight(80)
        if self.task_data and self.task_data.get("notes"):
            self.notes_edit.setPlainText(self.task_data["notes"])
        layout.addRow(_("Notes:"), self.notes_edit)

        # 按钮
        button_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addRow(button_box)

    def on_special_time_changed(self, text):
        """特殊时间选项改变时的处理"""
        current_data = self.special_time_combo.currentData()
        if current_data == "normal":
            self.reminder_time_edit.setEnabled(True)
        else:
            self.reminder_time_edit.setEnabled(False)

    def get_task_data(self):
        """获取任务数据"""
        # 根据特殊时间选项设置完成日期
        special_time_type = self.special_time_combo.currentData()
        due_date = self.due_date_edit.dateTime().toString("yyyy-MM-dd")

        if special_time_type == "no_deadline":
            due_date = "1999-09-09"
        elif special_time_type == "long_term":
            due_date = "1999-09-10"

        return {
            "category": self.category_combo.currentText(),
            "subcategory": self.subcategory_combo.currentText(),
            "description": self.description_edit.toPlainText().strip(),
            "due_date": due_date,
            "reminder_type": self.reminder_type_combo.currentText(),
            "reminder_time": self.reminder_time_edit.time().toString("HH:mm"),
            "status": self.status_combo.currentText(),
            "notes": self.notes_edit.toPlainText().strip(),
        }


class TasksManagerTab(QWidget):
    """任务管理器标签页"""

    def __init__(self, plugin_manager=None):
        super().__init__()
        self.plugin_manager = plugin_manager
        self.tasks_file = TASKS_FILE
        self.categories_file = CATEGORIES_FILE
        self.tasks_data = []
        self.remind_timers = {}
        self.notified_tasks = {}  # 记录已通知的任务，避免重复通知

        # 确保数据目录存在
        os.makedirs(DAILY_TASKS_DATA_DIR, exist_ok=True)

        self.init_ui()
        self.load_tasks()

        # 启动定时器检查提醒
        self.reminder_timer = QTimer()
        self.reminder_timer.timeout.connect(self.check_reminders)
        self.reminder_timer.start(10000)  # 每 10 秒检查一次

    def init_ui(self):
        """初始化界面"""
        main_layout = QVBoxLayout()
        self.setLayout(main_layout)

        # 顶部控制区域
        control_frame = QFrame()
        control_layout = QHBoxLayout()
        control_frame.setLayout(control_layout)

        # 筛选控件 - 使用独立的多选筛选组件
        # 传入 item_display_mapper 函数用于翻译显示文本
        self.category_filter = MultiSelectFilter(
            _("Category:"), DEFAULT_CATEGORIES.copy(), item_display_mapper=tr_category
        )
        self.category_filter.filterChanged.connect(self.apply_filters)
        control_layout.addWidget(self.category_filter)

        self.subcategory_filter = MultiSelectFilter(
            _("Subcategory:"), DEFAULT_SUBCATEGORIES.copy(), item_display_mapper=tr_category
        )
        self.subcategory_filter.filterChanged.connect(self.apply_filters)
        control_layout.addWidget(self.subcategory_filter)

        self.status_filter = MultiSelectFilter(
            _("Status:"), DEFAULT_STATUSES.copy(), item_display_mapper=tr_status
        )
        self.status_filter.filterChanged.connect(self.apply_filters)
        control_layout.addWidget(self.status_filter)

        # 加载用户自定义分类到筛选器
        self.load_categories_to_filters()

        control_layout.addStretch()

        # 排序控件
        control_layout.addWidget(QLabel(_("Sort:")))
        self.sort_combo = QComboBox()
        self.sort_combo.addItem(_("By Category"), "categories")
        self.sort_combo.addItem(_("By Due Date"), "due")
        self.sort_combo.currentIndexChanged.connect(self.apply_filters)
        control_layout.addWidget(self.sort_combo)

        # 倒序选项
        self.reverse_check = CustomCheckBox(_("Reverse"))
        self.reverse_check.stateChanged.connect(self.apply_filters)
        control_layout.addWidget(self.reverse_check)

        control_layout.addStretch()

        main_layout.addWidget(control_frame)

        # 第二行：操作按钮
        button_frame = QFrame()
        button_layout = QHBoxLayout()
        button_frame.setLayout(button_layout)

        self.add_btn = QPushButton(_("➕ Add Task"))
        self.add_btn.clicked.connect(lambda :self.add_task(True))
        button_layout.addWidget(self.add_btn)

        self.edit_btn = QPushButton(_("✏️ Edit Task"))
        self.edit_btn.clicked.connect(self.edit_task)
        button_layout.addWidget(self.edit_btn)

        self.delete_btn = QPushButton(_("🗑️ Delete Task"))
        self.delete_btn.clicked.connect(self.delete_task)
        button_layout.addWidget(self.delete_btn)

        self.complete_btn = QPushButton(_("✅ Mark Complete"))
        self.complete_btn.clicked.connect(self.mark_complete)
        button_layout.addWidget(self.complete_btn)

        button_layout.addStretch()

        main_layout.addWidget(button_frame)

        # 任务表格
        self.table = QTableWidget()
        # 表格列定义（在 init_ui 中定义以确保翻译生效）
        columns = [
            "ID",
            _("Category"),
            _("Subcategory"),
            _("Description"),
            _("Due Date"),
            _("Reminder"),
            _("Reminder Time"),
            _("Status"),
            _("Notes"),
        ]
        self.table.setColumnCount(len(columns))
        self.table.setHorizontalHeaderLabels(columns)

        # 设置表格属性
        self.table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch
        )
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.table.verticalHeader().setVisible(False)
        self.table.setMouseTracking(True)  # 启用鼠标追踪以支持 tooltip

        # 设置列宽
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)  # ID
        header.setSectionResizeMode(
            1, QHeaderView.ResizeMode.ResizeToContents
        )  # 任务大类
        header.setSectionResizeMode(
            2, QHeaderView.ResizeMode.ResizeToContents
        )  # 任务小类
        header.setSectionResizeMode(
            7, QHeaderView.ResizeMode.ResizeToContents
        )  # 完成状态
        header.setSectionResizeMode(
            4, QHeaderView.ResizeMode.ResizeToContents
        )  # 任务完成日期
        header.setSectionResizeMode(
            5, QHeaderView.ResizeMode.ResizeToContents
        )  # 提醒方式
        header.setSectionResizeMode(
            6, QHeaderView.ResizeMode.ResizeToContents
        )  # 提醒时间
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)  # 任务说明
        # 设置任务说明列的宽度至少大于300
        self.table.setColumnWidth(3, 300)
        header.setSectionResizeMode(8, QHeaderView.ResizeMode.Stretch)  # 备注

        # 连接双击信号到编辑函数
        self.table.cellDoubleClicked.connect(self.edit_task_on_doubleclick)

        main_layout.addWidget(self.table)

        # 底部状态栏
        status_frame = QFrame()
        status_layout = QHBoxLayout()
        status_frame.setLayout(status_layout)

        self.status_label = QLabel(_("Total: {} | Incomplete: {} | Completed: {} | In Progress: {} | Cancelled: {} | Not Started: {}").format(0, 0, 0, 0, 0, 0))
        status_layout.addWidget(self.status_label)

        status_layout.addStretch()

        self.refresh_btn = QPushButton(_("🔄 Refresh"))
        self.refresh_btn.clicked.connect(self.load_tasks)
        status_layout.addWidget(self.refresh_btn)

        main_layout.addWidget(status_frame)

        self.init_default_state()

    def init_default_state(self):
        self.reverse_check.setChecked(True)
        self.sort_combo.setCurrentIndex(0)
        
        # 设置完成状态筛选器默认不选择"已完成"
        default_statuses = [s for s in DEFAULT_STATUSES if s != "Completed"]
        self.status_filter.set_selected_items(default_statuses)

    def load_categories_to_filters(self):
        """加载分类到筛选器"""
        try:
            if os.path.exists(self.categories_file):
                with open(self.categories_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    categories = data["categories"]
                    subcategories = data["subcategories"]
                    # 更新筛选器的项目列表
                    self.category_filter.set_items(categories)
                    self.subcategory_filter.set_items(subcategories)
            else:
                print("Categories file not found. Using default categories.", self.categories_file)
        except Exception as e:
            print(f"加载分类失败：{e}")

    def save_categories(self):
        """保存分类到文件"""
        try:
            # 去重并保持原有顺序
            categories = list(dict.fromkeys(self.category_filter.items))
            subcategories = list(dict.fromkeys(self.subcategory_filter.items))
            
            data = {
                "categories": categories,
                "subcategories": subcategories,
            }
            with open(self.categories_file, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=4)
            
            # 更新内存中的列表为去重后的版本
            self.category_filter.items = categories
            self.subcategory_filter.items = subcategories
        except Exception as e:
            print(f"保存分类失败：{e}")

    def edit_task_on_doubleclick(self, row, column):
        """双击单元格时编辑任务"""
        self.edit_task()

    def load_tasks(self):
        """加载任务数据"""
        try:
            if os.path.exists(self.tasks_file):
                wb = openpyxl.load_workbook(self.tasks_file)
                ws = wb.active

                self.tasks_data = []
                for row_idx, row in enumerate(
                    ws.iter_rows(min_row=2, values_only=True), start=1
                ):
                    if row[0] is not None:  # ID 列不为空
                        task = {
                            "id": row[0],
                            "category": row[1] or "",
                            "subcategory": row[2] or "",
                            "description": row[3] or "",
                            "due_date": str(row[4]) if row[4] else "",
                            "reminder_type": row[5] or "",
                            "reminder_time": str(row[6]) if row[6] else "",
                            "status": row[7] or "",
                            "notes": row[8] or "",
                        }
                        # 处理旧数据中的日期时间格式，只保留日期部分
                        if task["due_date"]:
                            try:
                                # 如果是旧格式（包含时间），转换为只有日期
                                if " " in task["due_date"]:
                                    dt = datetime.strptime(
                                        task["due_date"], "%Y-%m-%d %H:%M"
                                    )
                                    task["due_date"] = dt.strftime("%Y-%m-%d")
                            except:
                                pass
                        self.tasks_data.append(task)

                wb.close()
            else:
                self.tasks_data = []
        except Exception as e:
            QMessageBox.critical(self, _("Error"), _("Failed to load tasks: {}").format(str(e)))
            self.tasks_data = []
        # 重置已通知记录，避免重启后不再提醒
        self.notified_tasks = {}
        # print("data", self.tasks_data)
        # # self.refresh_table()
        self.update_status()
        self.apply_filters()

    def save_tasks(self):
        """保存任务数据"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "每日任务"

            # 写入表头
            ws.append(
                [
                    "ID",
                    "任务大类",
                    "任务小类",
                    "任务说明",
                    "任务完成日期",
                    "提醒方式",
                    "提醒时间",
                    "完成状态",
                    "备注",
                ]
            )

            # 写入数据
            for task in self.tasks_data:
                ws.append(
                    [
                        task.get("id", ""),
                        task.get("category", ""),
                        task.get("subcategory", ""),
                        task.get("description", ""),
                        task.get("due_date", ""),
                        task.get("reminder_type", ""),
                        task.get("reminder_time", ""),
                        task.get("status", ""),
                        task.get("notes", ""),
                    ]
                )

            # 自动调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(self.tasks_file)
            wb.close()
        except Exception as e:
            exec_main_thread_callback(lambda : QMessageBox.critical(self, _("Error"), _("Failed to save tasks: {}").format(str(e))))

    def refresh_table(self, filtered_data=None):
        """刷新表格显示"""
        data_to_show = filtered_data if filtered_data is not None else self.tasks_data
        self.table.setRowCount(0)
        self.table.clearSpans()  # 清除之前的合并

        # 第一遍：插入所有行数据
        for task in data_to_show:
            row_position = self.table.rowCount()
            self.table.insertRow(row_position)

            self.table.setItem(
                row_position, 0, QTableWidgetItem(str(task.get("id", "")))
            )

            # 任务大类（先填充所有单元格，后续合并）
            category_item = QTableWidgetItem(task.get("category", ""))
            category_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row_position, 1, category_item)

            # 任务小类（先填充所有单元格，后续合并）
            subcategory_item = QTableWidgetItem(task.get("subcategory", ""))
            subcategory_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row_position, 2, subcategory_item)

            # 任务说明列添加 tooltip
            description_item = QTableWidgetItem(task.get("description", ""))
            description_item.setToolTip(task.get("description", ""))
            self.table.setItem(row_position, 3, description_item)

            # 处理特殊日期的显示
            due_date = task.get("due_date", "")
            if due_date.startswith("1999-"):
                due_date_display = ""
                for date_str, text, optionname in SPECIAL_DATES:
                    if due_date.startswith(date_str):
                        due_date_display = text
                        break
                assert due_date_display, f"Special date text {due_date} not found"
            else:
                due_date_display = due_date

            due_date_item = QTableWidgetItem(due_date_display)
            
            # 为日期列添加颜色提示（支持多层级红色系）
            urgency = TaskColorManager.get_date_urgency(due_date)
            if urgency != DateUrgency.NORMAL:
                bg_color, text_color = TaskColorManager.get_date_urgency_color(urgency)
                due_date_item.setBackground(bg_color)
                due_date_item.setForeground(text_color)
                # 添加提示文字
                urgency_text = TaskColorManager.get_urgency_display_text(urgency)
                if urgency_text:
                    due_date_item.setToolTip(urgency_text)
            
            due_date_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row_position, 4, due_date_item)
            self.table.setItem(
                row_position, 5, QTableWidgetItem(tr_reminder_type(task.get("reminder_type", "")))
            )
            self.table.setItem(
                row_position, 6, QTableWidgetItem(task.get("reminder_time", ""))
            )

            # 状态列根据状态设置颜色（状态颜色保持一致，不受日期影响）
            status = task.get("status", "")
            status_item = QTableWidgetItem(tr_status(status))
            
            # 只使用状态颜色，确保同一状态颜色一致
            bg_color, text_color = TaskColorManager.get_status_color(status)
            status_item.setBackground(bg_color)
            status_item.setForeground(text_color)
            # 设置文字居中
            status_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

            self.table.setItem(row_position, 7, status_item)

            # 备注列添加 tooltip
            notes_item = QTableWidgetItem(task.get("notes", ""))
            notes_item.setToolTip(task.get("notes", ""))
            self.table.setItem(row_position, 8, notes_item)

        # 第二遍：计算并设置合并
        if len(data_to_show) > 1:
            self._merge_category_columns(data_to_show)

    def _merge_category_columns(self, data_to_show):
        """计算并设置大类和小类列的合并"""
        n_rows = len(data_to_show)
        
        # 处理大类列（列1）
        i = 0
        while i < n_rows:
            current_category = data_to_show[i].get("category", "")
            # 查找连续相同的行数
            span_count = 1
            for j in range(i + 1, n_rows):
                if data_to_show[j].get("category", "") == current_category:
                    span_count += 1
                else:
                    break
            # 如果有多行相同，设置合并
            if span_count > 1:
                self.table.setSpan(i, 1, span_count, 1)
            i += span_count
        
        # 处理小类列（列2）
        i = 0
        while i < n_rows:
            current_subcategory = data_to_show[i].get("subcategory", "")
            # 查找连续相同的行数
            span_count = 1
            for j in range(i + 1, n_rows):
                if data_to_show[j].get("subcategory", "") == current_subcategory:
                    span_count += 1
                else:
                    break
            # 如果有多行相同，设置合并
            if span_count > 1:
                self.table.setSpan(i, 2, span_count, 1)
            i += span_count

    def apply_filters(self):
        """应用筛选条件"""
        # 获取选中的项目
        ALL = object()
        selected_categories = self.category_filter.get_selected() if not self.category_filter.is_select_all() else ALL
        selected_subcategories = self.subcategory_filter.get_selected() if not self.subcategory_filter.is_select_all() else ALL
        selected_statuses = self.status_filter.get_selected() if not self.status_filter.is_select_all() else ALL

        filtered = []
        for task in self.tasks_data:
            # 检查是否符合筛选条件（全选时不过滤）
            category_match = (task.get("category") in selected_categories) if selected_categories is not ALL else True
            subcategory_match = (task.get("subcategory") in selected_subcategories) if selected_subcategories is not ALL else True
            status_match = (task.get("status") in selected_statuses) if selected_statuses is not ALL else True

            if category_match and subcategory_match and status_match:
                filtered.append(task)

        # 应用排序
        sorted_data = self.sort_data(filtered)
        self.refresh_table(sorted_data)

    def sort_data(self, data):
        """对数据进行排序"""
        sort_key = self.sort_combo.currentData()
        reverse = self.reverse_check.isChecked()

        if sort_key == "categories":
            # 按类别排序：先按大类，再按小类
            sorted_data = sorted(
                data,
                key=lambda x: (x.get("category", ""), x.get("subcategory", "")),
                reverse=reverse,
            )
        elif sort_key == "due":
            # 按截止期限排序
            def parse_date(task):
                due_date = task.get("due_date", "")
                try:
                    if " " in due_date:
                        return datetime.strptime(due_date, "%Y-%m-%d %H:%M")
                    else:
                        return datetime.strptime(due_date, "%Y-%m-%d")
                except:
                    return datetime.max  # 无法解析的日期排到最后

            sorted_data = sorted(data, key=parse_date, reverse=reverse)
        else:
            sorted_data = data

        return sorted_data

    def add_task(self, show_dialog=True):
        """添加新任务"""
        if show_dialog:
            dialog = TaskDialog(self)
            if dialog.exec() == int(QDialog.DialogCode.Accepted):
                task_data = dialog.get_task_data()

                # 生成新 ID
                max_id = max([t.get("id", 0) for t in self.tasks_data], default=0)
                new_id = max_id + 1
                task_data["id"] = new_id

                category = task_data.get("category", "")
                if category:
                    self.category_filter.items.append(category)
                subcategory = task_data.get("subcategory", "")
                if subcategory:
                    self.subcategory_filter.items.append(subcategory)

                self.tasks_data.append(task_data)
                self.save_tasks()
                
                def save_and_load():
                    self.save_categories()
                    self.load_categories_to_filters()
                    self.load_tasks()
                exec_main_thread_callback(save_and_load)
                QMessageBox.information(self, _("Success"), _("Task added successfully!"))
        return True

    def edit_task(self):
        """编辑任务"""
        selected_rows = self.table.selectedItems()
        if not selected_rows:
            QMessageBox.warning(self, _("Warning"), _("Please select a task to edit!"))
            return

        row = selected_rows[0].row()
        item = self.table.item(row, 0)
        if not item:
            return
        task_id = int(item.text())

        # 查找任务数据
        task_data = None
        for task in self.tasks_data:
            if task.get("id") == task_id:
                task_data = task.copy()
                break

        if task_data:
            dialog = TaskDialog(self, task_data)
            if dialog.exec() == int(QDialog.DialogCode.Accepted):
                updated_data = dialog.get_task_data()
                updated_data["id"] = task_id

                # 更新数据
                for i, task in enumerate(self.tasks_data):
                    if task.get("id") == task_id:
                        self.tasks_data[i] = updated_data
                        break

                self.save_tasks()
                self.load_tasks()
                QMessageBox.information(self, _("Success"), _("Task updated successfully!"))

    def delete_task(self):
        """删除任务"""
        selected_rows = self.table.selectedItems()
        if not selected_rows:
            QMessageBox.warning(self, _("Warning"), _("Please select a task to delete!"))
            return

        reply = QMessageBox.question(
            self,
            _("Confirm Delete"),
            _("Are you sure you want to delete the selected task?"),
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )

        if reply == QMessageBox.StandardButton.Yes:
            row = selected_rows[0].row()
            item = self.table.item(row, 0)
            if not item:
                return
            task_id = int(item.text())

            # 删除数据
            self.tasks_data = [t for t in self.tasks_data if t.get("id") != task_id]
            self.save_tasks()
            self.load_tasks()
            QMessageBox.information(self, _("Success"), _("Task deleted successfully!"))

    def mark_complete(self):
        """标记任务为完成"""
        selected_rows = self.table.selectedItems()
        if not selected_rows:
            QMessageBox.warning(self, _("Warning"), _("Please select a task to mark as complete!"))
            return

        row = selected_rows[0].row()
        item = self.table.item(row, 0)
        if not item:
            return
        task_id = int(item.text())

        # 更新状态
        for i, task in enumerate(self.tasks_data):
            if task.get("id") == task_id:
                self.tasks_data[i]["status"] = "已完成"
                break

        self.save_tasks()
        self.load_tasks()
        QMessageBox.information(self, _("Success"), _("Task marked as complete!"))

    def check_reminders(self):
        """检查是否需要提醒"""
        current_time = datetime.now()
        current_date_str = current_time.strftime("%Y-%m-%d")
        current_time_str = current_time.strftime("%H:%M")

        for task in self.tasks_data:
            if task.get("status") == "已完成":
                continue

            reminder_type = task.get("reminder_type", "")
            reminder_time_str = task.get("reminder_time", "")

            if not reminder_time_str:
                continue

            # 检查是否到达提醒时间
            if reminder_time_str != current_time_str:
                continue

            # 根据提醒类型判断是否需要提醒
            should_remind = False
            task_key = f"{task.get('id')}_{current_date_str}"

            if reminder_type == "Daily":
                # 每天都提醒，只需要检查今天是否已经提醒过
                if task_key not in self.notified_tasks:
                    should_remind = True
                    self.notified_tasks[task_key] = True

            elif reminder_type == "On Due Date":
                # 只在完成日期当天提醒
                due_date = task.get("due_date", "")
                if due_date:
                    try:
                        # 兼容旧格式（包含时间）和新格式（只有日期）
                        if " " in due_date:
                            due_date_obj = datetime.strptime(due_date, "%Y-%m-%d %H:%M")
                        else:
                            due_date_obj = datetime.strptime(due_date, "%Y-%m-%d")
                        due_date_str = due_date_obj.strftime("%Y-%m-%d")
                        if (
                            due_date_str == current_date_str
                            and task_key not in self.notified_tasks
                        ):
                            should_remind = True
                            self.notified_tasks[task_key] = True
                    except:
                        pass

            elif reminder_type == "Once":
                # 只提醒一次
                if task_key not in self.notified_tasks:
                    should_remind = True
                    self.notified_tasks[task_key] = True

            if should_remind:
                self.show_system_notification(task)

        # 清理过期的通知记录（保留今天的记录）
        self.notified_tasks = {
            k: v for k, v in self.notified_tasks.items() if current_date_str in k
        }

    def show_system_notification(self, task):
        """显示系统通知"""
        title = _("📋 Task Reminder")
        message = f"【{task.get('category', '')} - {task.get('subcategory', '')}】\n{task.get('description', '')}\n\n{_('Due Date:')}{task.get('due_date', '')}"

        try:
            # 使用 plyer 发送系统通知
            plyer.notification.notify(
                title=title,
                message=message,
                app_name=_("Daily Task Reminder"),
                timeout=10,  # 通知显示 10 秒
            )
        except Exception as e:
            print(f"Notification failed: {e}")
            # 如果系统通知失败，回退到弹窗
            QMessageBox.information(self, title, message)

    def update_status(self):
        """更新状态栏"""
        total = len(self.tasks_data)
        completed = sum(1 for t in self.tasks_data if t.get("status") == "Completed")
        incomplete = total - completed
        in_progress = sum(1 for t in self.tasks_data if t.get("status") == "In Progress")
        cancelled = sum(1 for t in self.tasks_data if t.get("status") == "Cancelled")
        not_started = sum(1 for t in self.tasks_data if t.get("status") == "Not Started")

        self.status_label.setText(
            _("Total: {} | Incomplete: {} | Completed: {} | In Progress: {} | Cancelled: {} | Not Started: {}").format(
                total, incomplete, completed, in_progress, cancelled, not_started)
        )

    # ==================== API 接口方法（暴露给其他插件调用） ====================

    def get_all_enum_values_api(self) -> dict:
        """
        API: 获得全部可能的枚举值清单，当添加任务时不知道枚举值时，需要用该方法查询。

        Args:
            无

        Returns:
            dict[str, list[str]]: 过滤器名称列表，比如 {"category": ["类别1", "类别2"], "subcategory": ["小类1", "小类2"], "reminder_type": [...], "status": [...] }
        """
        return {
            "category": self.category_filter.items,
            "subcategory": self.subcategory_filter.items,
            "reminder_type": REMINDER_TYPES,
            "status": DEFAULT_STATUSES
        }
    
    def add_task_api(self, task_data: dict) -> int:
        """
        API: 添加任务，注意如果发现有的枚举值不知道该怎么填的时候，应当首先查询枚举值清单，不要随意填写。

        Args:
            task_data: 任务数据字典，包含以下字段：
                - category(str): 任务大类，需查阅枚举值清单，如果没有贴切的，就取名为 "其他"
                - subcategory(str): 任务小类，需查阅枚举值清单，如果没有贴切的，就取名为 "其他"
                - description(str): 任务说明，不能为空
                - due_date(str): 完成日期 (格式：yyyy-MM-dd)，不能为空
                - reminder_type(str): 提醒方式，需查阅枚举值清单，如果没有则设置为“其他”
                - reminder_time(str): 提醒时间 (格式：HH:mm), 如 "12:00"，如果无需设置则设置为当日"10:30"
                - status(str): 完成状态，需查阅枚举值清单，如果没有则设置为“其他”
                - notes(str): 备注

        Returns:
            int: 新任务的 ID，如果失败返回 -1
        """
        try:
            # 生成新 ID
            max_id = max([t.get("id", 0) for t in self.tasks_data], default=0)
            new_id = max_id + 1

            # 创建任务数据
            new_task = {
                "id": new_id,
                "category": task_data.get("category", ""),
                "subcategory": task_data.get("subcategory", ""),
                "description": task_data.get("description", ""),
                "due_date": task_data.get("due_date", ""),
                "reminder_type": task_data.get("reminder_type", "Once"),
                "reminder_time": task_data.get("reminder_time", ""),
                "status": task_data.get("status", "未开始"),
                "notes": task_data.get("notes", ""),
            }

            # 添加到数据列表
            self.tasks_data.append(new_task)
            self.save_tasks()

            # 更新分类
            category = new_task.get("category", "")
            if category and category not in self.category_filter.items:
                self.category_filter.items.append(category)
            subcategory = new_task.get("subcategory", "")
            if subcategory and subcategory not in self.subcategory_filter.items:
                self.subcategory_filter.items.append(subcategory)

            self.save_categories()
            self.load_categories_to_filters()
            self.load_tasks()

            return new_id
        except Exception as e:
            print(f"[DailyTasks] 添加任务失败：{e}")
            import traceback
            traceback.print_exc()
            return -1

    def delete_task_api(self, task_id: int) -> bool:
        """
        API: 删除任务

        Args:
            task_id: 任务 ID

        Returns:
            bool: 是否删除成功
        """
        try:
            # 查找任务
            task_exists = any(t.get("id") == task_id for t in self.tasks_data)
            if not task_exists:
                return False

            # 删除任务
            self.tasks_data = [t for t in self.tasks_data if t.get("id") != task_id]
            self.save_tasks()
            self.load_tasks()
            return True
        except Exception as e:
            print(f"[DailyTasks] 删除任务失败：{e}")
            return False
    
    def convert_task(self, task: dict) -> dict:
        """
        替换任务的特殊日期为文字，例如将1999年9月10日替换为“长期”，并且返回新的task dict
        """
        task = task.copy()
        for date_str, text, optionname in SPECIAL_DATES:
            if date_str in task.get("due_date", ""):
                task["due_date"] = text
                break
        return task
    
    def get_todo_tasks_api(self) -> list:
        """
        API: 获取状态非“完成”的任务列表

        Returns:
            list: 状态非“完成”的任务列表
        """
        try:
            return [self.convert_task(t) for t in self.tasks_data if t.get("status") != "已完成"]
        except Exception as e:
            print(f"[DailyTasks] 获取未完成任务列表失败：{e}")
            return []

    def get_tasks_api(self, filters: dict = None) -> list:
        """
        API: 获取任务列表

        Args:
            filters: 过滤条件（可选），支持：
                - category: 任务大类（字符串或列表）
                - subcategory: 任务小类（字符串或列表）
                - status: 完成状态（字符串或列表）
                - due_date_from: 开始日期 (格式：yyyy-MM-dd)
                - due_date_to: 结束日期 (格式：yyyy-MM-dd)

        Returns:
            list: 任务列表
        """
        try:
            if not filters:
                return self.tasks_data.copy()

            filtered = []
            for task in self.tasks_data:
                # 检查大类
                if "category" in filters:
                    filter_cat = filters["category"]
                    if isinstance(filter_cat, str):
                        filter_cat = [filter_cat]
                    if task.get("category") not in filter_cat:
                        continue

                # 检查小类
                if "subcategory" in filters:
                    filter_subcat = filters["subcategory"]
                    if isinstance(filter_subcat, str):
                        filter_subcat = [filter_subcat]
                    if task.get("subcategory") not in filter_subcat:
                        continue

                # 检查状态
                if "status" in filters:
                    filter_status = filters["status"]
                    if isinstance(filter_status, str):
                        filter_status = [filter_status]
                    if task.get("status") not in filter_status:
                        continue

                # 检查日期范围
                if "due_date_from" in filters or "due_date_to" in filters:
                    due_date = task.get("due_date", "")
                    if due_date:
                        if "due_date_from" in filters:
                            if due_date < filters["due_date_from"]:
                                continue
                        if "due_date_to" in filters:
                            if due_date > filters["due_date_to"]:
                                continue

                filtered.append(task)

            return filtered
        except Exception as e:
            print(f"[DailyTasks] 获取任务列表失败：{e}")
            return []

    def get_task_by_id_api(self, task_id: int) -> dict:
        """
        API: 根据 ID 获取任务

        Args:
            task_id: 任务 ID

        Returns:
            dict: 任务数据，不存在返回 None
        """
        try:
            for task in self.tasks_data:
                if task.get("id") == task_id:
                    return task.copy()
            return None
        except Exception as e:
            print(f"[DailyTasks] 获取任务失败：{e}")
            return None

    def update_task_api(self, task_id: int, task_data: dict) -> bool:
        """
        API: 更新任务

        Args:
            task_id: 任务 ID
            task_data: 新的任务数据

        Returns:
            bool: 是否更新成功
        """
        try:
            # 查找任务
            for i, task in enumerate(self.tasks_data):
                if task.get("id") == task_id:
                    # 保留 ID
                    updated_task = task.copy()
                    updated_task.update(task_data)
                    updated_task["id"] = task_id
                    self.tasks_data[i] = updated_task
                    self.save_tasks()
                    self.load_tasks()
                    return True
            return False
        except Exception as e:
            print(f"[DailyTasks] 更新任务失败：{e}")
            return False

    def filter_tasks_by_date_api(
        self, start_date: str, end_date: str = None
    ) -> list:
        """
        API: 按日期过滤任务

        Args:
            start_date: 开始日期 (格式：yyyy-MM-dd)
            end_date: 结束日期（可选，格式：yyyy-MM-dd），不提供则只过滤当天的

        Returns:
            list: 任务列表
        """
        try:
            if end_date is None:
                end_date = start_date

            filtered = []
            for task in self.tasks_data:
                due_date = task.get("due_date", "")
                if due_date and start_date <= due_date <= end_date:
                    filtered.append(task)

            return filtered
        except Exception as e:
            print(f"[DailyTasks] 按日期过滤任务失败：{e}")
            return []

    def mark_task_complete_api(self, task_id: int) -> bool:
        """
        API: 标记任务为完成

        Args:
            task_id: 任务 ID

        Returns:
            bool: 是否成功
        """
        try:
            for i, task in enumerate(self.tasks_data):
                if task.get("id") == task_id:
                    self.tasks_data[i]["status"] = "已完成"
                    self.save_tasks()
                    self.load_tasks()
                    return True
            return False
        except Exception as e:
            print(f"[DailyTasks] 标记任务完成失败：{e}")
            return False


# ==================== 插件入口函数 ====================


def load_plugin(plugin_manager: PluginManager):
    """
    插件加载入口函数

    Args:
        plugin_manager: 插件管理器实例

    Returns:
        dict: 包含插件组件的字典
    """
    print("[DailyTasks] 正在加载每日任务提醒器插件...")

    # 创建标签页实例
    tasks_tab = TasksManagerTab(plugin_manager=plugin_manager)

    # 注册暴露的方法到全局域
    plugin_manager.register_method(
        "daily_tasks", "get_all_enum_values", tasks_tab.get_all_enum_values_api,
        extra_data={"enable_mcp": True}
    )
    plugin_manager.register_method(
        "daily_tasks", "add_task", tasks_tab.add_task_api,
        extra_data={"enable_mcp": True}
    )
    plugin_manager.register_method(
        "daily_tasks", "delete_task", tasks_tab.delete_task_api
    )
    plugin_manager.register_method(
        "daily_tasks", "get_tasks", tasks_tab.get_tasks_api,
        extra_data={"enable_mcp": True}
    )
    plugin_manager.register_method(
        "daily_tasks", "get_todo_tasks", tasks_tab.get_todo_tasks_api,
        extra_data={"enable_mcp": True}
    )
    plugin_manager.register_method(
        "daily_tasks", "get_task_by_id", tasks_tab.get_task_by_id_api
    )
    plugin_manager.register_method(
        "daily_tasks", "update_task", tasks_tab.update_task_api
    )
    plugin_manager.register_method(
        "daily_tasks", "filter_tasks_by_date", tasks_tab.filter_tasks_by_date_api
    )
    plugin_manager.register_method(
        "daily_tasks", "mark_task_complete", tasks_tab.mark_task_complete_api
    )

    # 添加到标签页（由插件管理器统一管理）
    plugin_manager.add_plugin_tab("daily_tasks", _("📋 Task Management"), tasks_tab, position=1)

    print("[DailyTasks] 每日任务提醒器插件加载完成")
    return {"tab": tasks_tab, "namespace": "daily_tasks"}


def unload_plugin(plugin_manager):
    """
    插件卸载回调

    Args:
        plugin_manager: 插件管理器实例
    """
    print("[DailyTasks] 正在卸载每日任务提醒器插件...")
    # 清理资源、保存状态等
    print("[DailyTasks] 每日任务提醒器插件卸载完成")
