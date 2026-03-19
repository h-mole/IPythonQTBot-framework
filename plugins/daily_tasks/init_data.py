"""
初始化每日任务插件的数据文件
创建 Excel 文件并设置表头
"""

import os
import openpyxl

# 数据目录
DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
TASKS_FILE = os.path.join(DATA_DIR, "daily_tasks.xlsx")

def init_excel_file():
    """初始化 Excel 文件"""
    # 确保目录存在
    os.makedirs(DATA_DIR, exist_ok=True)
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "每日任务"
    
    # 写入表头
    headers = [
        "ID",
        "任务大类",
        "任务小类",
        "任务说明",
        "任务完成日期",
        "提醒方式",
        "提醒时间",
        "完成状态",
        "备注"
    ]
    ws.append(headers)
    
    # 设置列宽
    column_widths = [8, 12, 12, 40, 15, 12, 12, 12, 30]
    for i, width in enumerate(column_widths, start=1):
        col_letter = openpyxl.utils.get_column_letter(i)
        ws.column_dimensions[col_letter].width = width
    
    # 设置表头样式
    from openpyxl.styles import Font, Alignment, PatternFill
    
    header_font = Font(bold=True, size=11)
    header_alignment = Alignment(horizontal='center', vertical='center')
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill
    
    # 保存文件
    wb.save(TASKS_FILE)
    wb.close()
    
    print(f"✓ 已创建 Excel 文件：{TASKS_FILE}")

if __name__ == "__main__":
    init_excel_file()
    print("初始化完成！")
